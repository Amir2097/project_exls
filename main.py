import glob
import os
import openpyxl
import pandas as pd


def folders(dirxlsx=None):

    if dirxlsx is None:
        ref_dir = os.path.abspath(os.curdir)
    else:
        ref_dir = dirxlsx

    return ref_dir


def extract_all_files():
    files_list = []
    print(folders())
    for file in sorted(glob.glob(f'{folders()}/Склад[0-9]*.xls*')):
        print(file)
        files_list.append(os.path.join(folders(), file))

    print(files_list)
    return files_list


def read_xlsx():
    for i in extract_all_files():
        wookbook = openpyxl.load_workbook(i)
        # Define variable to read the active sheet:
        worksheet = wookbook.active
        # Iterate the loop to read the cell values
        for ii in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                print(col[ii].value, end="\t\t")
            print('\n')


if __name__ == '__main__':
    read_xlsx()
    # extract_all_files()
